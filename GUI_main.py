#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
This script is the main script used to produce a GUI which should help for
cell segmentation. This script can only read .nd2 files containing the 
images of cells, especially it displays for each recorded positions (field
of view) the pictures in the time axis.

The script opens first a window which allows you to load an nd2 file and
to load or create an hdf file. The hdf file contains all the masks, so if
it is the first the user segments an nd2 file, a new one should be created.
And it can be then loaded for later use.  Along with new hdf file, created
by the name entered by the user (say filename), it creates three other hdf 
files (filename_predicted.h5, filename_thresholded.h5 and 
filename_segmented.h5) these contain all the steps of the NN to get to the 
segmented picture. 

After the first window is finished a second one opens, where at each time 
index, three pictures 
are displayed the t-1 picture, the t picture (current frame which can be
edited) and the t+1 picture. Using the arrows one can navigate through time.
On top of the picture, there is always a mask which is displayed, if no cells
are present in the mask then the mask is blank and the user does not see it. 
If one wants to hand anmotate the pictures, one can just start to draw on the
picture using the different functions (New Cell, Add Region, Brush, Eraser,
Save Mask, ...) and the informations will be saved in the mask overlayed on 
top of the pictures. 

If one wants to segment using a neural network, one can press the
corresponding button (Launch CNN) and select the time range and 
the field of views on which the neural network is applied.

Once the neural network has finished predicting, there are still no visible
masks, but on the field of views and time indices where the NN has been
applied, the threshold and segment buttons are enabled. By checking these
two buttons one can either display the thresholded image of the prediction or
display the segmentation of the thresholded prediction.

At this stage, one can correct the segmentation of the prediction using
the functions (New Cell, Add Region, etc..) by selecting the Segment
checkbox and then save them using the Save Seg button.
If the user is happy with the segmentation, the Cell Correspondance button 
can be clicked. Until then, the cells get random numbers attributed by
the segmentation algorithm. In order to keep track of the cell through time,
the same cells should have the same number between two different time pictures.
This can be (with some errors) achieved by the Cell Correspondance button,
which tries to attribute the same number to corresponding cells in time.
After that, the final mask is saved and it is always visible when you go on
the corresponding picture. This mask can also be corrected using the 
usual buttons (because the Cell Correspondance makes also mistakes). 

"""
import sys
import numpy as np

# For writing excel files
from openpyxl import load_workbook
from openpyxl import Workbook

# Import everything for the Graphical User Interface from the PyQt5 library.
from PyQt5.QtWidgets import (QApplication, QMainWindow, QDialog, 
    QMessageBox, QPushButton, QCheckBox, QAction, QStatusBar, QLabel)
from PyQt5 import QtGui

#Import from matplotlib to use it to display the pictures and masks.
from matplotlib.backends.qt_compat import QtWidgets
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar

#append all the paths where the modules are stored. Such that this script
#looks into all of these folders when importing modules.
sys.path.append("./unet")
sys.path.append("./disk")
sys.path.append("./icons")
sys.path.append("./init")
sys.path.append("./misc")

#Import all the other python files
#this file handles the interaction with the disk, so loading/saving images
#and masks and it also runs the neural network.
import InteractionDisk_temp as nd

#this file contains a dialog window that takes two integers as entry to swap
#two cell values
import ExchangeCellValues as ecv

#this file contains a dialog window which is opened before the main program
#and allows to load the nd2 and hdf files by browsing through the computer.
import DialogFileBrowser as dfb

#this file contains a window that opens to change the value of one cell. It 
#is opened as soon as the user presses with the left click on a specific cell.
import ChangeOneCellValue as cocv

#this file contains a dialog window to browse for the excel file where
#all the extracted information on the fluoerscence is written. Or to create a 
#new excel file by typing a name in the text box. It is thought to have one 
#excel file per field of view.
import DialogDataBrowser as ddb

#this file contains a dialog window where a time range and the field of views
#can be selected to then launch a prediction of the neural network on
#a specific range of pictures.
import LaunchBatchPrediction as lbp

#this file initializes all the buttons present in the gui, sets the shortcuts
#to these buttons and also connect the buttons to the function that are 
#triggered when the buttons are pressed.
import InitButtons

#this file contains the layout of the main window so it justs puts the buttons
#and the pictures at the desired position in the main window.
import InitLayout

# PlotCanvas for fast plotting
from PlotCanvas import PlotCanvas
    

class NavigationToolbar(NavigationToolbar):
    """This is the standard matplotlib toolbar but only the buttons
    that are of interest for this gui are loaded. These buttons allow 
    to zoom into the pictures/masks and to navigate in the zoomed picture. 
    A Home button can be used to set the view back to the original view.
    """
    toolitems = [t for t in NavigationToolbar.toolitems if
                 t[0] in ('Home', 'Pan', 'Zoom','Back', 'Forward')]
      
class App(QMainWindow):
    """This class creates the main window.
    """

    def __init__(self, nd2pathstr, hdfpathstr, newhdfstr):
        super().__init__()
        self.title = 'YeaZ 1.0'

        # all these ids are integers which are used to set a connection between
        # the button and the function that this button calls.
        # There are three of them because it happens that one can trigger three 
        # different functions with one button.        
        self.id = 0
        self.id2 = 0
        self.id3 = 0 

        self.reader = nd.Reader(hdfpathstr, newhdfstr, nd2pathstr)
        
        # these variables are used to create/read/load the excel file used
        # to write the fluorescence values extracted. For each field of view,
        # the user will be asked each time to create a new xls file for the 
        # field of view or to load an existing field of view (this is the role
        # of the boolean variable)
        self.xlsfilename = ''
        self.nd2path = nd2pathstr
        self.FlagFluoExtraction = False
        
        # Set the indices for the time axis and the field of view index. These
        # indices represent everywhere the current picture (the one that can be
        # edited, i.e. the time t frame)
        self.Tindex = 0
        self.FOVindex = 0
        
        # loading the first images of the cells from the nd2 file
        self.currentframe = self.reader.LoadOneImage(self.Tindex,self.FOVindex)
        
        # check if the t+1 time frame exists, avoid failure if there is only
        # one picture in the folder/nd2 file
        if self.Tindex+1 < self.reader.sizet:
            self.nextframe = self.reader.LoadOneImage(self.Tindex+1, self.FOVindex)
        else:
            self.nextframe = np.zeros([self.reader.sizey, self.reader.sizex])
        
        self.previousframe = np.zeros([self.reader.sizey, self.reader.sizex])

        # loading the first masks from the hdf5 file
        self.mask_curr = self.reader.LoadMask(self.Tindex, self.FOVindex)
        self.mask_previous = np.zeros([self.reader.sizey, self.reader.sizex])
        
        # check if the t+1 mask exists, avoid failure if there is only
        # one mask in the hdf file
        if self.Tindex+1 < self.reader.sizet:
            self.mask_next = self.reader.LoadMask(self.Tindex+1, self.FOVindex)
        else:
            self.mask_next = np.zeros([self.reader.sizey, self.reader.sizex])
        
        # creates a list of all the buttons, which will then be used in order
        # to disable all the other buttons at once when one button/function
        # is pressed/used in the gui.
        self.buttonlist = []
        
        # setting buttons as attributes
        # the shortcuts for the buttons, the functions to which they are
        # connected to,... are all set up in the ButtonInit file which is called
        # in the self.initUI() method below.
        self.button_newcell = QPushButton("New cell")
        self.buttonlist.append(self.button_newcell)
        
        self.button_add_region = QPushButton("Add region")
        self.buttonlist.append(self.button_add_region)
        
        self.button_savemask = QPushButton("Save Mask")
        self.buttonlist.append(self.button_savemask)
        
        self.button_drawmouse = QPushButton('Brush')
        self.buttonlist.append(self.button_drawmouse)
        
        self.button_eraser = QPushButton('Eraser')
        self.buttonlist.append(self.button_eraser)
        
        self.button_exval = QPushButton('Exchange Cell Values')
        self.buttonlist.append(self.button_exval)
        
        self.button_showval = QCheckBox('Show Cell Values')
        self.buttonlist.append(self.button_showval)
        
        self.button_hidemask = QCheckBox('Hide Mask')
        self.buttonlist.append(self.button_hidemask)
        
        self.button_nextframe = QPushButton("Next Time Frame")
        self.buttonlist.append(self.button_nextframe)
        
        self.button_previousframe = QPushButton("Previous Time Frame")
        self.buttonlist.append(self.button_previousframe)
        
        self.button_cnn = QPushButton('Launch CNN')
        self.buttonlist.append(self.button_cnn)
        
        self.button_threshold = QCheckBox('Threshold prediction')
        self.buttonlist.append(self.button_threshold)
        
        self.button_segment = QCheckBox('Segment')
        self.buttonlist.append(self.button_segment)
                 
        self.button_cellcorespondance = QPushButton('Tracking')
        self.buttonlist.append(self.button_cellcorespondance)
        
        self.button_changecellvalue = QPushButton('Change cell value')
        self.buttonlist.append(self.button_changecellvalue)        
        
        self.button_extractfluorescence = QPushButton('Extract Fluorescence')
        self.buttonlist.append(self.button_extractfluorescence)
        
        self.button_hide_show = QPushButton('CNN')
        self.buttonlist.append(self.button_hide_show)
        
        self.initUI()


    def initUI(self):
        """Initializing the widgets contained in the window. 
        Especially, it creates the widget to plot the 
        pictures/masks by creating an object of the PlotCanvas class self.m. 
        Every interaction with the masks or the pictures (loading new
        frames/editing the frames/masks) occurs through this class.
        
        This method initializes all the buttons with the InitButtons file.
        It connects the buttons to the functions that they should trigger,
        it sets the shortcuts to the buttons, a tool tip, 
        eventually a message on the status bar when the user hovers 
        over the button, etc..
        
        This function also sets all the layout in the InitLayout file. It 
        takes and places the widgets (buttons, canvas, toolbar).

        The function initializes a Menu Bar to have a menu which can be 
        improved later on.
        It sets a toolbar of the matplotlib library and hides it. But it allows
        to connect to the functions of this toolbar through "homemade" 
        QPushButtons instead of the ones provided by matplotlib.
        Finally, it sets a StatusBar which displays some text to describe
        the use of some buttons, or to show that the program is working on 
        something (running the neural network, loading frames, etc...)
        
        After all this has been initialized, the program is ready to be used.
        """
        self._main = QtWidgets.QWidget()
        self.setCentralWidget(self._main)

        # Here our canvas is created where using matplotlib, 
        # one can plot data to display the pictures and masks.
        self.m = PlotCanvas(self)
        
        # Initialize all the buttons that are needed and the functions that are 
        # connected when the buttons are triggered.
        InitButtons.Init(self)
        InitLayout.Init(self)
        
        # MENU, TOOLBAR AND STATUS BAR
        # creates a menu just in case, some other functions can be added later
        # in this menu.
        menubar = self.menuBar()
        self.fileMenu = menubar.addMenu('File')   
        self.saveactionmenu = QAction('Save')
        self.fileMenu.addAction(self.saveactionmenu)
        self.saveactionmenu.triggered.connect(self.ButtonSaveMask)
        
        # hide the toolbar and instead of the original buttons of matplotlib,
        # QPushbuttons are used and are connected to the functions of the toolbar
        # it is than easier to interact with these buttons (for example to 
        # to disable them and so on..)
        self.Nvgtlbar = NavigationToolbar(self.m, self)
        self.addToolBar(self.Nvgtlbar)
        self.Nvgtlbar.hide()
        
        # creates a status bar with user instructions
        self.statusBar = QStatusBar()
        self.setStatusBar(self.statusBar)
        self.statusBarText = QLabel()
        self.statusBar.addWidget(self.statusBarText)
                
        self.show()
                
    
    def mousePressEvent(self, QMouseEvent):
        """this function is implemented just to have the QLineButtons of the 
        change time index button, setthreshold button and the setsegmentation
        button out of focus when the user clicks somewhere
        on the gui. (to unfocus the buttons)
        """
        self.button_timeindex.clearFocus()
        if self.button_SetThreshold.isEnabled():
            self.button_SetThreshold.clearFocus()
            
        if self.button_SetSegmentation.isEnabled():
            self.button_SetSegmentation.clearFocus()
        
        
# -----------------------------------------------------------------------------
# FUNCTIONS LINKED TO NAVIGATION
# connect the functions of the toolbar to our custom QPushbuttons.
    def ZoomTlbar(self):
        """The button_zoom is connected to the zoom function of the toolbar 
        already present in the matplotlib library.
        
        Depending on the buttons that are active or checked, when the zoom 
        function is used, it does not disable all the buttons.
        
        If the segment and threshold button are not checked or used
        when the zoom button is clicked, it disables all the button
        using self.Disable which disables everything except the button passed
        in argument (in this case button_zoom).
        
        If the zoom button is used while the segment button is checked,
        it disables all the buttons (1st elif) except the segment button
        but once it is finished (so the zoom button becomes unchecked) 
        then it enables only the editing buttons (as long as the segment
        button is still checked) such as New Cell, Add Region, Eraser, 
        Brush,etc.. and the other toolbar buttons (3rd elif)
        
        
        If the zoom button is clicked while the threshold button is checked,
        it disables all the button except the threshold button (2nd elif).
        Once the zoom button is unchecked, it enables the toolbar buttons
        (4th elif)
        In any other case, it just enables all the buttons again.
        """
        self.Nvgtlbar.zoom()
        
        if (self.button_zoom.isChecked() and not(self.button_segment.isChecked() 
            or self.button_threshold.isChecked())):
            self.Disable(self.button_zoom)
            
        elif self.button_zoom.isChecked() and self.button_segment.isChecked():
            self.Disable(self.button_zoom)
            self.button_segment.setEnabled(True)
            
        elif self.button_zoom.isChecked() and self.button_threshold.isChecked():
            self.Disable(self.button_zoom)
            self.button_threshold.setEnabled(True)
            
        elif self.button_zoom.isChecked() == False and self.button_segment.isChecked():
            self.button_pan.setEnabled(True)
            self.button_home.setEnabled(True)
            self.button_back.setEnabled(True)
            self.button_forward.setEnabled(True)
            self.EnableCorrectionsButtons()
            
        elif self.button_zoom.isChecked() == False and self.button_threshold.isChecked():
            self.button_pan.setEnabled(True)
            self.button_home.setEnabled(True)
            self.button_back.setEnabled(True)
            self.button_forward.setEnabled(True)
            
        else:
            self.Enable(self.button_zoom)
        
        
    def HomeTlbar(self):
        """
        connects the home button to the home function of the matplotlib
        toolbar. It sets the view to the original view (no zoom)
        """
        self.Nvgtlbar.home()

            
    def BackTlbar(self):
        """
        It calls the back function of the matplotlib toolbar which sets the 
        view to the previous one (if the user does several zooms/pans, 
        this button allows to go back in the "history of views")
        """
        self.Nvgtlbar.back()

        
        
    def ForwardTlbar(self):
        """
        It calls the forward function of the matplotlib toolbar which sets the 
        view to the next one (if the user does several zooms/pans, 
        this button allows to go forward in the "history of views"
        """
        self.Nvgtlbar.forward()

    
    def PanTlbar(self):
        """The button_pan is connected to the pan function of the toolbar 
        already present in the matplotlib library.
        
        Depending on the buttons that are active or checked, when the pan 
        function is used, it does not disable all the buttons.
        
        If the segment and threshold button are not checked or used
        when the pan button is clicked, it disables all the button
        using self.Disable which disables everything except the button passed
        in argument (in this case button_pan).
        
        If the pan button is used while the segment button is checked,
        it disables all the buttons (1st elif) except the segment button
        but once it is finished (so the zoom button becomes unchecked) 
        then it enables only the editing buttons (as long as the segment
        button is still checked) such as New Cell, Add Region, Eraser, 
        Brush,etc.. and the other toolbar buttons (3rd elif)
        
        
        If the pan button is clicked while the threshold button is checked,
        it disables all the button except the threshold button (2nd elif).
        Once the pan button is unchecked, it enables the toolbar buttons
        (4th elif)
        In any other case, it just enables all the buttons again.
        """

        self.Nvgtlbar.pan()

        if (self.button_pan.isChecked() and not(self.button_segment.isChecked()
            or self.button_threshold.isChecked())):
            self.Disable(self.button_pan)
            
        elif self.button_pan.isChecked() and self.button_segment.isChecked():
            self.Disable(self.button_pan)
            self.button_segment.setEnabled(True)
            
        elif self.button_pan.isChecked() and self.button_threshold.isChecked():
            self.Disable(self.button_pan)
            self.button_threshold.setEnabled(True)
            
        elif not(self.button_pan.isChecked()) and self.button_segment.isChecked():
            
            self.button_zoom.setEnabled(True)
            self.button_home.setEnabled(True)
            self.button_back.setEnabled(True)
            self.button_forward.setEnabled(True)
            
            self.EnableCorrectionsButtons()
            
        elif not(self.button_pan.isChecked()) and self.button_threshold.isChecked():
            self.button_zoom.setEnabled(True)
            self.button_home.setEnabled(True)
            self.button_back.setEnabled(True)
            self.button_forward.setEnabled(True)
        
        else:
            self.Enable(self.button_pan)


# -----------------------------------------------------------------------------
# EXTRACTING FLUORESCENCE
            
    def ButtonFluo(self):
        """This function is called everytime the Extract Fluorescence button is 
        clicked (self.button_extractfluorescence). 
        
        self.FlagFluoExtraction is boolean which is True when the path to the 
        excel file has already been loaded into self.xlsfilename.
        This pathname changes for each field of view as it is thought to have
        one xls file per field of view. 
        So at the beginning and each time the user changes field of view,
        self.FlagFluoExtraction is set to False.
        
        When it is set to False, this function calls a dialog window where
        the user is asked to load an already existing xls file for the current
        field of view or to give a name to create a new xls file for
        the current field of view. (self.Dialogxls)
        
        If it set to true, it means that self.xlsfilename contains the path
        to the xls file for the current field of view and it is directly given
        to the function that writes the fluorescence into the xls file.
        (self.ExtractFluo)
        """        
        if self.FlagFluoExtraction:
            self.ExtractFluo(self.xlsfilename)
        else:
            self.DialogXls()
        
        
    def DialogXls(self):
        """This function creates a dialog window which gives two options to the
        user either to load an existing xls file or to give a new name in order
        to create a new xls file. 
        """
        # creates the window
        dwind = ddb.FileBrowser()

        # this test is True if the user presses ok in the dialog window, if the 
        # user presses cancels it returns and nothing happens.
        if dwind.exec_():
            xlsname = dwind.xlsname
            newxlsname = dwind.newxlsentry.text()

            # if the string containing the filepath to an existing xls file 
            # is not empty then it calls directly the function to write the 
            # data into this existing xls file and sets self.xlsfilename
            if xlsname:
                self.xlsfilename = xlsname
                self.ExtractFluo(xlsname)
                
            # if xlsname is empty then it creates a new pathfilename and puts
            # the new created xls file into the folder where nd2 is located.
            # the string containing the nd2 namepath is split
            else:
                xlsname = ''
                templist = self.nd2path.split('/')
                
                for k in range(0, len(templist)-1):
                    
                    xlsname = xlsname + templist[k] + '/'
                # this is the new path/filename
                xlsname = xlsname + newxlsname + '.xlsx'
                self.xlsfilename = xlsname
                
                self.CreateXls(xlsname)
                self.ExtractFluo(xlsname)
                
            # this flag is set to true, for the current field of view each
            # time extract fluorescence is clicked it writes in the file located
            # at self.xlsfilename.
            self.FlagFluoExtraction = True
            
        else:
            return


    def CreateXls(self, xlsfilename):
        """In case there is no xls file existing, here a new one is created 
        and prepared. For each channel a new sheet is created.
        In the first row for each sheet, the time indices are written t = 0,
        t = 1, etc... but only every third column. Because in the row below,
        three values are extracted 'Total intensity', 'Area' and 'Variance'
        at each time index. So three columns for each time index are needed,
        for the three data points. 
        The first column is left empty (starting from third row) because
        the cell numbers will be written in there.
        """
        
        book = Workbook()
        nbrchannels = self.reader.sizec
        
        for i in range(0,nbrchannels):
            sheetname = self.reader.channel_names[i]
            # creates a sheet with the name of the corresponding channel.
            if i == 0:
                sheet = book.active
                sheet.title = sheetname
            else:
                sheet = book.create_sheet(sheetname)
            sheet.cell(1,1, 'Cell Number / Time axis')
            sheet.cell(2,1, 'labels')
            timeaxissize = self.reader.sizet
            # start writing the time index at column 1, column 0 is reserved for
            # cell numbers.
            timecolindex = 2
            
            for t in range(1,timeaxissize+1):
                sheet.cell(1,timecolindex).value = 't = {}'.format(t-1)
                sheet.cell(2,timecolindex).value = 'Total Intensity'
                sheet.cell(2,timecolindex+1).value = 'Total Area'
                sheet.cell(2,timecolindex+2).value = 'Mean Intensity'
                sheet.cell(2,timecolindex+3).value =  'Variance'
                # updates the index, where the next time index should be written
                timecolindex = timecolindex + 4
              
        try:
            book.save(xlsfilename)
        except TypeError:
            QMessageBox.critical(self, "Error", "TypeError encountered. \
                                 Make sure you have openpyxl version 3.0.1 \
                                 installed. If the problem persists contact \
                                 the developers.")

        
    def ExtractFluo(self, xlsfilename):
        """This is the function that takes as argument the filepath to the xls
        file and writes in the file.
        It iterates over the different channels (or the sheets of the file,
        each channel has one sheet.), and reads the image corresponding
        to the time, field of view and channel index. It reads the already
        existing file and makes a copy in which the data will be written in it.
        
        The first step of calculating the data is to iterate through each
        cell/segment of the mask (so each cell is a submatrix of one value
        in the matrix of the mask).
        For each of these value /cell, the area is extracted as being
        the number of pixels corresponding to this cell/value. 
        (it is known from the microscope settings how to convert
        the pixel in area).
        The total intensity is just the value of the pixel and it is added over
        all the pixels corresonding to the cell/value.
        The mean is then calculated as being the total intensity divided by
        the number of pixels (which here is equal to the area also).
        With the mean it is then possible to calculate the variance of the 
        signal for one cell/value.
        
        Then, it is checked if the value of the cell (cell number) already
        exists in the first column, if it already exists it continues to
        find the column corresponding to the time index where the values
        should be written. It sets the flag to True such that it does not
        write the cell as new one and adds it at the end of the column
        
        If the value is not found in the cell number column (new cell or
        first time writing in the file), the flag is False, thus it adds the 
        cell number at the end of the column.
        It then saves the xls file.
        """
        self.Disable(self.button_extractfluorescence)
        self.WriteStatusBar('Extracting the fluorescence...')
        
        # opens the file to read it.
        book = load_workbook(self.xlsfilename)


        # iterate over all the channels, so over all the sheets in the file
        for channel in range(0, self.reader.sizec):
            # loads the picture corresponding to the channel, time index and fov
            image = self.reader.LoadImageChannel(self.Tindex, self.FOVindex, channel)
            
            # loads the sheet to read out corresponding to the current channel
            sheet = book.worksheets[channel]
            
            
            # this index contains the value of the maximum number of rows in the
            # file, it is used to append at the end the cell number column a new
            # cell/value, and it is updated each time a new cell is added.
            tempidx = sheet.max_row
           
            # np.unique(array) returns an array which contains all the value
            # that appear in self.m.plotmask, so it returns every cell value
            # including the background (value 0) present in self.m.plotmask
            for val in np.unique(self.m.plotmask):
                
                # Skip background
                if val == 0:
                    continue
                
                # Calculate stats
                area = (self.m.plotmask == val).sum()
                tot_intensity = image[self.m.plotmask == val].sum()
                mean = tot_intensity/area
                var = np.var(image[self.m.plotmask==val])

                # if flag is false it means that the cell number
                # corresponding to val is not present in the xls file, first
                # column.
                flag = False
                
                # iterate over all the rows
                for row in range(sheet.max_row+1):
                     # test if in the first column 0, the number of the cell
                     # is already present
                     # if sheet.cell_value(row,0) == str(val):
                     if sheet.cell(row = row+1, column = 1).value == str(val):
                         
                         # if is present, the column corresponding to the
                         # current time index is by iterating over the cols.
                         for col in range(sheet.max_column+1):
                             # test if it is the right column
                             # if sheet.cell_value(0, col) == 't = {}'.format(self.Tindex):
                             if sheet.cell(row = 1, column = col+1).value == 't = {}'.format(self.Tindex):
                                 # write in the xls file at the row, col coord
                                 sheet.cell(row+1, col+1, str(tot_intensity))
                                 sheet.cell(row+1, col+2, str(area))
                                 sheet.cell(row+1,col+3, str(mean))
                                 sheet.cell(row+1, col+4, str(var))
                                 book.save(xlsfilename)

                                 # the flag is set to True so that it does
                                 # not execute the code where the cell is
                                 # added in the xls file in a new row.
                                 flag = True
                                 
                if not flag:
                # this lines are executed if a new cell is detected or if
                # if it is the first time to write in the file.
                    for col in range(sheet.max_column+1):
                        if sheet.cell(row = 1, column =  col+1).value == 't = {}'.format(self.Tindex):
                            # it write the cell value/cell number in the
                            # column
                            sheet.cell(tempidx+1,1, str(val))
                            
                            # writes the data extracted before
                            sheet.cell(tempidx+1,col+1,str(tot_intensity))
                            sheet.cell(tempidx+1, col+2, str(area))
                            sheet.cell(tempidx+1, col+3, str(mean))
                            sheet.cell(tempidx+1, col+4, str(var))
                            # it updates the number of rows as a new cell
                            # has been added, so there is one more row.
                            tempidx = tempidx + 1
                            book.save(xlsfilename)

        self.Enable(self.button_extractfluorescence)
        self.ClearStatusBar()


# -----------------------------------------------------------------------------
# NEURAL NETWORK
    def ShowHideCNNbuttons(self):
        
        """hide and show the buttons corresponding to the neural network.
            this function is called by the button CNN which is hidden. But
            if activated in the InitLayout.py then you can have a button
            which hides the CNN buttons (which are now on the normal also
            hidden...).
        """
        if self.button_hide_show.isChecked():
            self.button_cnn.setVisible(True)
            self.button_segment.setVisible(True)
            self.button_savesegmask.setVisible(True)
            self.button_threshold.setVisible(True)
            self.button_SetThreshold.setVisible(True)
            self.button_savethresholdmask.setVisible(True)
            self.button_SetSegmentation.setVisible(True)

        else:
            self.button_cnn.setVisible(False)
            self.button_segment.setVisible(False)
            self.button_savesegmask.setVisible(False)
            self.button_threshold.setVisible(False)
            self.button_SetThreshold.setVisible(False)
            self.button_savethresholdmask.setVisible(False)
            self.button_SetSegmentation.setVisible(False)
            

    def LaunchBatchPrediction(self):
        """This function is called whenever the button Launch CNN is pressed.
        It allows to run the neural network over a time range and selected
        field of views.
        
        It creates a dialog window with two entries, that define the time range
        and a list where the user can select the desired fields of view.
        
        Once it reads all the value, it calls the neural network function
        inside of self.PredThreshSeg and it does the prediction of the neural
        network, thresholds this prediction and then segments it.
        """
        # creates a dialog window from the LaunchBatchPrediction.py file
        dlg = lbp.CustomDialog(self)
        
        # this if tests if the user pressed 'ok' in the dialog window
        if dlg.exec_() == QDialog.Accepted:
            # it tests if the user has entered some values
            # if not it ignores and returns.
            if not (dlg.entry1.text()!= '' and dlg.entry2.text() != ''):
                QMessageBox.critical(self, "Error", "No Time Specified")
                return 
            
            # reads out the entry given by the user and converts the index
            # to integers
            time_value1 = int(dlg.entry1.text())
            time_value2 = int(dlg.entry2.text())
    
            # it tests if the first value is smaller or equal such that
            # time_value1 is the lower range of the time range
            # and time_value2 the upper boundary of the range.
            if time_value1 > time_value2 :
                QMessageBox.critical(self, "Error", 'Invalid Time Constraints')
                return
            
            # displays that the neural network is running
            self.WriteStatusBar('Running the neural network...')
    
            #it iterates in the list of the user-selected fields 
            #of view, to return the corresponding index, the function
            #dlg.listfov.row(item) is used which gives an integer
            if len(dlg.listfov.selectedItems())==0:
                QMessageBox.critical(self, "Error", "No FOV Selected")
            
            for item in dlg.listfov.selectedItems():
                #iterates over the time indices in the range
                for t in range(time_value1, time_value2+1):                    
                    #calls the neural network for time t and selected
                    #fov
                    if dlg.entry_threshold.text() !=  '':
                        thr_val = float(dlg.entry_threshold.text())
                    else:
                        thr_val = None
                    if dlg.entry_segmentation.text() != '':
                        seg_val = int(dlg.entry_segmentation.text())
                    else:
                        seg_val = 10
                    self.PredThreshSeg(t, dlg.listfov.row(item), thr_val, seg_val)
                    
                    # if tracker has been checked then apply it
                    if dlg.tracking_checkbox.isChecked():
                        if t != 0:
                            temp_mask = self.reader.CellCorrespondance(t,dlg.listfov.row(item))
                            self.reader.SaveMask(t,dlg.listfov.row(item), temp_mask)
                        
                        else:
                            temp_mask = self.reader.LoadSeg(t, dlg.listfov.row(item))
                            self.reader.SaveMask(t,dlg.listfov.row(item), temp_mask)
            
            self.ReloadThreeMasks()
               
            #once it has iterated over all the fov, the message in 
            #the status bar is cleared and the buttons are enabled.
            self.ClearStatusBar()
            self.EnableCNNButtons()
   
    
    def PredThreshSeg(self, timeindex, fovindex, thr_val, seg_val):
          """
          This function is called in the LaunchBatchPrediction function.
          This function calls the neural network function in the
          InteractionDisk.py file and then thresholds the result
          of the prediction, saves this thresholded prediction.
          Then it segments the thresholded prediction and saves the
          segmentation. 
          """
          self.reader.LaunchPrediction(timeindex, fovindex)
          self.m.ThresholdMask = self.reader.ThresholdPred(thr_val, timeindex,fovindex)
          self.reader.SaveThresholdMask(timeindex, fovindex, self.m.ThresholdMask)
          self.m.SegmentedMask = self.reader.Segment(seg_val, timeindex,fovindex)
          self.reader.SaveSegMask(timeindex, fovindex, self.m.SegmentedMask)
    
    
    def LaunchPrediction(self):
        """This function is not used in the gui, but it can be used to launch
        the prediction of one picture, with no thresholding and no segmentation
        """
        if not(self.reader.TestPredExisting(self.Tindex, self.FOVindex)):
            self.WriteStatusBar('Running the neural network...')
            self.Disable(self.button_cnn)
            self.reader.LaunchPrediction(self.Tindex, self.FOVindex)
            
            self.Enable(self.button_cnn)
            
            self.button_cnn.setEnabled(False)
            self.button_threshold.setEnabled(True)
            self.button_segment.setEnabled(True)
            self.button_cellcorespondance.setEnabled(True)
            self.ClearStatusBar()
        
    
    def SelectChannel(self, index):
        """This function is called when the button to select different channels
        is used. From the displayed list in the button, the chosen index
        corresponnds to the same index in the list of channels from the reader.
        So, it sets the default channel with the new index (called index below)
        """
        self.reader.default_channel = index
        # update the pictures using the same function as the one used to 
        # change the fields of view.
        self.ChangeFOV()
        

    def SelectFov(self, index):
        """This function is called when the button containing the list of 
        fields od view is used.
        The index correspondds to the field of view selected in the list.
        """
        # mask is automatically saved.
        self.reader.SaveMask(self.Tindex, self.FOVindex, self.m.plotmask)
        self.FOVindex = index    
        
        # it updates the fov in the plot with the new index.
        self.ChangeFOV()
        
        # the flag of the fluorescence extraction is set to False (such that
        # if the user extracts fluorescence data in the new field of  view,
        # there is a dialog box asking to select the corresponding xls file
        # for this field of view. IF there is no data sheet for this fov, the
        # user can enter a new name to make a new file.)
        self.FlagFluoExtraction = False
        
        
    def ChangeFOV(self):
        """
        it changes the fov or channel according to the choice of the user
        and it updates the plot shown and it initializes the new fov/channel
        at t=0 by default.
        """
        
        self.Tindex = 0
        
        # load the image and mask for the current plot
        self.m.currpicture = self.reader.LoadOneImage(self.Tindex,self.FOVindex)
        self.m.plotmask = self.reader.LoadMask(self.Tindex,self.FOVindex)
        
        # sets the image and the mask to 0 for the previous plot
        self.m.prevpicture = np.zeros([self.reader.sizey, self.reader.sizex], dtype = np.uint16)
        self.m.prevplotmask = np.zeros([self.reader.sizey, self.reader.sizex], dtype = np.uint16)
        
        # load the image and the mask for the next plot, check if it exists
        if self.Tindex+1 < self.reader.sizet:
            self.m.nextpicture = self.reader.LoadOneImage(self.Tindex+1, self.FOVindex)
            self.m.nextplotmask = self.reader.LoadMask(self.Tindex+1, self.FOVindex)
            
            # enables the next frame button in case it was disabled when the 
            # fov/channel was changed
            self.button_nextframe.setEnabled(True)
        else:
            self.m.nextpicture = np.zeros([self.reader.sizey, self.reader.sizex], dtype = np.uint16)
            self.m.nextplotmask =  np.zeros([self.reader.sizey, self.reader.sizex], dtype = np.uint16)
            
            # disables the next frame button if the mask or the picture
            # does not exist.
            self.button_nextframe.setEnabled(False)
            
        # once the images and masks are loaded into the variables, they are 
        # displaye in the gui.
        self.m.UpdatePlots()
        
        # disables the previous frame button in case it was active before 
        # changing fov/channel.
        self.button_previousframe.setEnabled(False)
        
        # updates the title of the plots to display the right time indices
        # aboves the plots.
        self.UpdateTitleSubplots()
            
        # if the button to hide the mask was checked before changing fov/channel,
        # it hides the mask again.
        if self.button_hidemask.isChecked():
            self.m.HideMask()
        
        # the button to set the time index is also set to 0/default again.
        self.button_timeindex.setText('')
        # enables the neural network buttons if there is already an 
        # existing prediction for the current image.
        self.EnableCNNButtons()
        
        
    def ReloadThreeMasks(self):
        """
        A function which replots all the masks at the current time and fov 
        indices. Needed after the batch prediction is completed to display
        the result of the NN.
        """
        
        if self.Tindex >= 0 and self.Tindex <= self.reader.sizet-1:
            if self.Tindex == 0:
                self.button_nextframe.setEnabled(True)
                
                if self.Tindex < self.reader.sizet-1:
                    self.m.nextplotmask = self.reader.LoadMask(self.Tindex+1, self.FOVindex)
                else:
                    np.zeros([self.reader.sizey, self.reader.sizex], dtype = np.uint16)
                
                self.m.plotmask = self.reader.LoadMask(self.Tindex, self.FOVindex)
                self.m.prevplotmask = np.zeros([self.reader.sizey, self.reader.sizex], dtype = np.uint16)
                self.m.UpdatePlots()
                self.button_previousframe.setEnabled(False)
                
                
            elif self.Tindex == self.reader.sizet-1:
                self.button_previousframe.setEnabled(True)
                self.m.prevplotmask = self.reader.LoadMask(self.Tindex-1, self.FOVindex)
                self.m.plotmask = self.reader.LoadMask(self.Tindex, self.FOVindex)
                self.m.nextplotmask =  np.zeros([self.reader.sizey, self.reader.sizex], dtype = np.uint16)
                self.m.UpdatePlots()
                self.button_nextframe.setEnabled(False)
                
            else:
                self.button_nextframe.setEnabled(True)
                self.button_previousframe.setEnabled(True)
                self.m.prevplotmask = self.reader.LoadMask(self.Tindex-1, self.FOVindex)
                self.m.plotmask = self.reader.LoadMask(self.Tindex, self.FOVindex)              
                self.m.nextplotmask = self.reader.LoadMask(self.Tindex+1, self.FOVindex)
                self.m.UpdatePlots()
            
            self.UpdateTitleSubplots()
                        
            if self.button_hidemask.isChecked():
                self.m.HideMask()
            self.EnableCNNButtons()
        
        else:
            return
        
    
    def ChangeTimeFrame(self):
        """This funcion is called whenever the user gives a new time index, 
        to jump to the new given index, once "enter" button is pressed.
        """
        
        # it reads out the text in the button and converts it to an int.
        newtimeindex = int(self.button_timeindex.text())
        if newtimeindex >= 0 and newtimeindex <= self.reader.sizet-1:
            self.reader.SaveMask(self.Tindex, self.FOVindex, self.m.plotmask)
            
            self.Tindex = newtimeindex
            
            if self.Tindex == 0:
                self.button_nextframe.setEnabled(True)
                self.m.nextpicture = self.reader.LoadOneImage(self.Tindex+1,self.FOVindex)
                self.m.nextplotmask = self.reader.LoadMask(self.Tindex+1, self.FOVindex)
                
                self.m.currpicture = self.reader.LoadOneImage(self.Tindex, self.FOVindex)
                self.m.plotmask = self.reader.LoadMask(self.Tindex, self.FOVindex)
                
                self.m.prevpicture = np.zeros([self.reader.sizey, self.reader.sizex], 
                                              dtype = np.uint16)
                self.m.prevplotmask = np.zeros([self.reader.sizey, self.reader.sizex], 
                                               dtype = np.uint16)
    
                self.m.UpdatePlots()
                self.button_previousframe.setEnabled(False)
                
            elif self.Tindex == self.reader.sizet-1:
                self.button_previousframe.setEnabled(True)
                self.m.prevpicture = self.reader.LoadOneImage(self.Tindex-1, self.FOVindex)
                self.m.prevplotmask = self.reader.LoadMask(self.Tindex-1, self.FOVindex)
                   
                self.m.currpicture = self.reader.LoadOneImage(self.Tindex, self.FOVindex)
                self.m.plotmask = self.reader.LoadMask(self.Tindex, self.FOVindex)
                  
                self.m.nextpicture =  np.zeros([self.reader.sizey, self.reader.sizex], 
                                               dtype = np.uint16)
                self.m.nextplotmask =  np.zeros([self.reader.sizey, self.reader.sizex], 
                                                dtype = np.uint16)
                
                self.m.UpdatePlots()
                self.button_nextframe.setEnabled(False)
                
            else:
                self.button_nextframe.setEnabled(True)
                self.button_previousframe.setEnabled(True)
                self.m.prevpicture = self.reader.LoadOneImage(self.Tindex-1, self.FOVindex)
                self.m.prevplotmask = self.reader.LoadMask(self.Tindex-1, self.FOVindex)
                   
                self.m.currpicture = self.reader.LoadOneImage(self.Tindex, self.FOVindex)
                self.m.plotmask = self.reader.LoadMask(self.Tindex, self.FOVindex)              
                  
                self.m.nextpicture = self.reader.LoadOneImage(self.Tindex+1,self.FOVindex)
                self.m.nextplotmask = self.reader.LoadMask(self.Tindex+1, self.FOVindex)
                self.m.UpdatePlots()
            
            self.UpdateTitleSubplots()
            self.button_timeindex.clearFocus()
            self.button_timeindex.setText(str(self.Tindex)+'/'+str(self.reader.sizet-1))
            
            if self.button_hidemask.isChecked():
                self.m.HideMask()
            self.EnableCNNButtons()
        
        else:
            self.button_timeindex.clearFocus()
            return
        
     
    def CellCorrespActivation(self):
            self.Disable(self.button_cellcorespondance)
            self.WriteStatusBar('Doing the cell correspondance')

            if self.Tindex != 0:
                self.m.plotmask = self.reader.CellCorrespondance(self.Tindex, self.FOVindex)
                self.m.updatedata()
            else:
                self.m.plotmask = self.reader.LoadSeg(self.Tindex, self.FOVindex)
                self.m.updatedata()

            self.Enable(self.button_cellcorespondance)
            self.button_cellcorespondance.setChecked(False)
            self.ClearStatusBar()
        
        
    def SegmentBoxCheck(self):
        if self.button_segment.isChecked():
            self.Disable(self.button_segment)
            self.EnableCorrectionsButtons()
            self.m.SegmentedMask = self.reader.LoadSeg(self.Tindex, self.FOVindex)
            self.m.tempplotmask = self.m.plotmask.copy()
            self.m.plotmask = self.m.SegmentedMask.copy()
            self.m.currmask.set_data((self.m.SegmentedMask%10 + 1)*(self.m.SegmentedMask != 0))
            self.m.ax.draw_artist(self.m.currplot)
            self.m.ax.draw_artist(self.m.currmask)
            self.m.update()
            self.m.flush_events()
            
            # update the graph
            self.button_SetSegmentation.setEnabled(True)
            self.button_savesegmask.setEnabled(True)
            
        else:
            self.m.SegmentedMask = self.m.plotmask.copy()
            self.m.plotmask = self.m.tempplotmask.copy()
            self.m.updatedata()
            self.button_SetSegmentation.setEnabled(False)
            self.button_savesegmask.setEnabled(False)
            self.Enable(self.button_segment)
    
    
    def SegmentThresholdedPredMask(self):
        # update the plots to display the segmentation view
        segparamvalue = int(self.button_SetSegmentation.text())
        self.m.plotmask = self.reader.Segment(segparamvalue, self.Tindex,self.FOVindex)
        self.m.currmask.set_data((self.m.plotmask%10 + 1)*(self.m.plotmask != 0))
        self.m.ax.draw_artist(self.m.currplot)
        self.m.ax.draw_artist(self.m.currmask)
        self.m.update()
        self.m.flush_events()
    
    
    def ButtonSaveSegMask(self):
        """saves the segmented mask
        """
        self.reader.SaveSegMask(self.Tindex, self.FOVindex, self.m.plotmask)
    
        
    def ThresholdBoxCheck(self):
        """if the buttons is checked it shows the thresholded version of the 
        prediction, if it is not available it justs displays a null array.
        The buttons for the setting a threshold a value and to save it are then
        activated once this button is enabled.
        """
        if self.button_threshold.isChecked():
            self.Disable(self.button_threshold)
            self.m.ThresholdMask = self.reader.LoadThreshold(self.Tindex, self.FOVindex)
            
            self.m.currmask.set_data(self.m.ThresholdMask)
            self.m.ax.draw_artist(self.m.currplot)
            self.m.ax.draw_artist(self.m.currmask)
            self.m.update()
            self.m.flush_events()
            
            self.button_SetThreshold.setEnabled(True)
            self.button_savethresholdmask.setEnabled(True)
            
        else:
            self.m.updatedata()
            self.button_SetThreshold.setEnabled(False)
            self.button_savethresholdmask.setEnabled(False)
            self.Enable(self.button_threshold)


    def ThresholdPrediction(self):
        # update the plots to display the thresholded view
        thresholdvalue = float(self.button_SetThreshold.text())
        
        self.m.ThresholdMask = self.reader.ThresholdPred(
            thresholdvalue, 
            self.Tindex,self.FOVindex)
        
        self.m.currmask.set_data(self.m.ThresholdMask)
        self.m.ax.draw_artist(self.m.currplot)
        self.m.ax.draw_artist(self.m.currmask)
        self.m.update()
        self.m.flush_events()
      
        
    def ButtonSaveThresholdMask(self):
        """saves the thresholed mask
        """
        self.reader.SaveThresholdMask(self.Tindex, self.FOVindex, self.m.ThresholdMask)

        
    def ChangePreviousFrame(self):
         """This function is called when the previous frame buttons is pressed 
         and it tests if the buttons is enabled and if so it calls the
         BackwardTime() function. It should avoid the let the user do multiple 
         clicks and that the function is then called afterwards several times,
         once the frames and masks of the current time index have been loaded.
         """
         if self.button_previousframe.isEnabled():
            self.button_previousframe.setEnabled(False)
            self.BackwardTime()
            if self.Tindex >0:
                self.button_previousframe.setEnabled(True)
         else:
             return
             
             
    def ChangeNextFrame(self):
        """This function is called when the next frame buttons is pressed 
        and it tests if the buttons is enabled and if so it calls the
        ForwardTime() function. It should avoid the let the user do multiple 
        clicks and that the function is then called afterwards several times,
        once the frames and masks of the current time index have been loaded.
        """
        if self.button_nextframe.isEnabled():
            self.button_nextframe.setEnabled(False)
            self.ForwardTime()
            if self.Tindex + 1 < self.reader.sizet:
                self.button_nextframe.setEnabled(True)
                
        else:
            return

        
    def ForwardTime(self):
        """This function switches the frame in forward time index. And it tests
        several conditions if t == lastTimeIndex-1, because then the next frame
        button has to be disabled. It also tests if the show value of cells
        button and hidemask are active in order to hide/show the mask or to 
        show the cell values.
        """
        # the t frame is defined as the currently shown frame on the display.
        # If the button "Next time frame" is pressed, this function is called
        self.WriteStatusBar('Loading the next frame...')
        self.Disable(self.button_nextframe)

        if self.Tindex + 1 < self.reader.sizet - 1 :
            self.reader.SaveMask(self.Tindex, self.FOVindex, self.m.plotmask)
            
            self.m.prevpicture = self.m.currpicture.copy()
            self.m.prevplotmask = self.m.plotmask.copy()
            
            self.m.currpicture = self.m.nextpicture.copy()
            self.m.plotmask = self.m.nextplotmask.copy()
            
            self.m.nextpicture = self.reader.LoadOneImage(self.Tindex+2, self.FOVindex)
            self.m.nextplotmask = self.reader.LoadMask(self.Tindex+2, self.FOVindex)
            self.m.UpdatePlots()

            if self.Tindex + 1 == 1:
                self.button_previousframe.setEnabled(True)
                
        else:
            self.reader.SaveMask(self.Tindex, self.FOVindex, self.m.plotmask)
        
            self.m.prevpicture = self.m.currpicture.copy()
            self.m.prevplotmask = self.m.plotmask.copy()
            self.m.currpicture = self.m.nextpicture.copy()
            self.m.plotmask = self.m.nextplotmask.copy()
            self.m.nextpicture = np.zeros([self.reader.sizey, self.reader.sizex], 
                                          dtype = np.uint16)
            self.m.nextplotmask = np.zeros([self.reader.sizey,self.reader.sizex], 
                                           dtype = np.uint16)
            self.m.UpdatePlots()

            self.button_nextframe.setEnabled(False)

        self.Tindex = self.Tindex+1
        self.UpdateTitleSubplots()
        
        if self.button_hidemask.isChecked():
            self.m.HideMask()

        self.Enable(self.button_nextframe)
        self.ClearStatusBar()
        self.button_timeindex.setText(str(self.Tindex)+'/'+str(self.reader.sizet-1))

    
    def BackwardTime(self):
        """This function switches the frame in backward time index. And it 
        several conditions if t == 1, because then the button previous frame has to
        be disabled. It also tests if the show value of cells button and 
        hidemask are active in order to hide/show the mask or to show the cell
        values.
        """
        # the t frame is defined as the currently shown frame on the display.
        # If the button "Previous time frame" is pressed, this function is called
        self.WriteStatusBar('Loading the previous frame...')
        self.Disable(self.button_previousframe)
        
        self.reader.SaveMask(self.Tindex, self.FOVindex, self.m.plotmask)

        self.m.nextpicture = self.m.currpicture.copy()
        self.m.nextplotmask = self.m.plotmask.copy()
        self.m.currpicture = self.m.prevpicture.copy()
        self.m.plotmask = self.m.prevplotmask.copy()
            
        if self.Tindex == 1:
            self.m.prevpicture = np.zeros([self.reader.sizey, self.reader.sizex], dtype = np.uint16)
            self.m.prevplotmask = np.zeros([self.reader.sizey, self.reader.sizex], dtype = np.uint16)
            self.button_previousframe.setEnabled(False)
            
        else:
            self.m.prevpicture = self.reader.LoadOneImage(self.Tindex-2, self.FOVindex)
            self.m.prevplotmask = self.reader.LoadMask(self.Tindex-2, self.FOVindex)

        self.m.UpdatePlots()
        if self.Tindex-1 == self.reader.sizet-2:
            self.button_nextframe.setEnabled(True)            
        
        if self.button_hidemask.isChecked():
            self.m.HideMask()
        
        self.Tindex -= 1
        self.UpdateTitleSubplots()
            
        self.Enable(self.button_previousframe)
        
        if self.Tindex > 0:
            self.button_previousframe.setEnabled(True)          
            
        self.ClearStatusBar()
        self.button_timeindex.setText(str(self.Tindex)+'/' + str(self.reader.sizet-1))


# -----------------------------------------------------------------------------
# MANUAL MASK CORRECTIONS
            
    def ChangeOneValue(self):
        """This function is called when the button Change cell value is
        clicked. It displays the instructions on the status bar.
        And if the user clicks in the graph where the current mask is displayed
        it connects the event of the click (meaning that user has clicked on
        one cell) to the function self.DialogBoxChangeOneValue. 
        This function will then replaces the cell selected by the user with
        the click with a new value entered by the user.
        """
        
        # displaying the instructions on the statusbar
        self.WriteStatusBar((
            'Select one cell using the left click '
             'and then enter the desired value in the dialog box.'))

        # disables all the buttons
        self.Disable(self.button_changecellvalue)
        
        # connects the event "press mouse button" in the matplotlib plot 
        # (picture) to the function self.DialogBoxChangeOneValue
        self.id = self.m.mpl_connect('button_press_event', self.DialogBoxChangeOneValue)
        
        
    def DialogBoxChangeOneValue(self, event):
        """This function is called when the user after the user has selected
        the button Change cell value and clicked in the picture to select
        the desired cell to change.
        
        It first deconnects the mouse click event in matplotlib with this 
        function to not generate any other dialog window.
        
        It then tests if the click is inside the matplotlib plot (if it is
        outside it equals to None) and if it is the current and editable plot
        (the one in the middle of the gui, self.m.ax)
        
        If is true, then it sets the coordinates to int. and creates a dialog
        window where the user is asked to type a value to set it to the cell.
        
        If the user presses ok, it tests if the entry is valid (>0 and not 
        empty) and looks for the old cell value and replaces it. And then
        it updates the plot such that the result of the change can be seen.
        """
        # the function is disconnected from the matplotlib event.
        self.m.mpl_disconnect(self.id)
        
        # test if the button is a left click and if the coordinates
        # chosen by the user click is inside of the current matplotlib plot
        # which is given by self.m.ax
        if (event.button == 1 
            and (event.xdata != None and event.ydata != None) 
            and self.m.ax == event.inaxes):
            newx = int(event.xdata)
            newy = int(event.ydata)
            
            # creates a dialog window
            dlg = cocv.CustomDialog(self)
            
            #if the user presses 'ok' in the dialog window it executes the code
            #else it does nothing
            if dlg.exec_():
                #it tests that the user has entered some value, that it is not
                #empty and that it is equal or bigger to 0.
                if dlg.entry1.text() != '' and int(dlg.entry1.text()) >= 0:
                    #reads the new value to set and converts it from str to int
                    value = int(dlg.entry1.text())
                    
                    # self.m.plotmask[newy, newx] the value selected by the user
                    # self.m.plotmask == self.m.plotmask[newy, newx]
                    # gives the coordinates where it is equal to the value
                    # selected by the user. And it replaces it with the new
                    # value.
                    self.m.plotmask[self.m.plotmask == self.m.plotmask[newy,newx]] = value
                    
                    # updates the plot to see the modification.
                    self.m.updatedata()
                    
        self.Enable(self.button_changecellvalue)
        self.button_changecellvalue.setChecked(False)
        self.m.ShowCellNumbers()
        
        
    def DialogBoxECV(self, s):
        """This functions creates from the ExchangeCellValues.py file a 
        window which takes two integer entries and then swaps the cells having
        the given integer values.
        """
        # creates a dialog window from the ExchangeCellValues.py file
        dlg = ecv.CustomDialog(self)
        
        # if the user presses 'ok', it executes the code
        if dlg.exec_():

            # it tests if both value to be swapped are not empty.
            if dlg.entry1.text()!= '' and dlg.entry2.text() != '':
                
                # reads out the values and converts it into integers.
                value1 = int(dlg.entry1.text())
                value2 = int(dlg.entry2.text())
                
                # calls the function which does the swap
                self.m.ExchangeCellValue(value1,value2)
                self.m.ShowCellNumbers()
                
        else:
            return


    def MouseDraw(self):
        """
        This function is called whenever the brush or the eraser button is
        pressed. On the first press event it calls the self.m.OneClick, which
        tests whether it is a right click or a left click. If it a right click
        it assigns the value of the pixel which has been right clicked
        to self.cellval, meaning that the next drawn pixels will be set to this
        value.
        If it is left clicked, then it draws a 3x3 square with the current
        value of self.cellval.
        If after left clicking you drag the mouse, then you start drawing 
        using the mouse and it stops once you release the left click.
        
        Same for the eraser button, it sets directly the value of self.cellval
        to 0.
        """
        if self.button_drawmouse.isChecked():
            self.WriteStatusBar(('Draw using the brush, right click to select '
                                 'the cell to draw.'))
            self.Disable(self.button_drawmouse)
            
            self.m.tempmask = self.m.plotmask.copy()
            
            self.id2 = self.m.mpl_connect('button_press_event', self.m.OneClick)
            self.id = self.m.mpl_connect('motion_notify_event', self.m.PaintBrush)
            self.id3 = self.m.mpl_connect('button_release_event', self.m.ReleaseClick)

            pixmap = QtGui.QPixmap('./icons/brush2.png')
            cursor = QtGui.QCursor(pixmap, -1,-1)
            QApplication.setOverrideCursor(cursor)
        
        elif self.button_eraser.isChecked():
            self.WriteStatusBar('Erasing by setting the values to 0.')
            self.Disable(self.button_eraser)
            
            self.m.tempmask = self.m.plotmask.copy()
            
            self.m.cellval = 0
            self.id2 = self.m.mpl_connect('button_press_event', self.m.OneClick)
            self.id = self.m.mpl_connect('motion_notify_event', self.m.PaintBrush)
            self.id3 = self.m.mpl_connect('button_release_event', self.m.ReleaseClick)
            
            pixmap = QtGui.QPixmap('./icons/eraser.png')
            cursor = QtGui.QCursor(pixmap, -1, -1)
            QApplication.setOverrideCursor(cursor)
            
        else:
            self.m.mpl_disconnect(self.id3)
            self.m.mpl_disconnect(self.id2)
            self.m.mpl_disconnect(self.id)
            QApplication.restoreOverrideCursor()
            self.Enable(self.button_drawmouse)
            self.Enable(self.button_eraser)
            
            self.ClearStatusBar()
            
            
    def UpdateTitleSubplots(self):
        """This function updates the title of the plots according to the 
        current time index. So it called whenever a frame or a fov is changed.
        """
        if self.Tindex == 0:
            self.m.titlecurr.set_text('Time index {}'.format(self.Tindex))
            self.m.titleprev.set_text('No frame {}'.format(''))
            self.m.titlenext.set_text('Next time index {}'.format(self.Tindex+1))
            self.m.draw()
            
        elif self.Tindex == self.reader.sizet-1:
            self.m.titlecurr.set_text('Time index {}'.format(self.Tindex))
            self.m.titleprev.set_text('Previous time index {}'.format(self.Tindex-1))
            self.m.titlenext.set_text('No frame {}'.format(''))            
            self.m.draw()
            
        else:
            self.m.titlecurr.set_text('Time index {}'.format(self.Tindex))
            self.m.titleprev.set_text('Previous time index {}'.format(self.Tindex-1))
            self.m.titlenext.set_text('Next time index {}'.format(self.Tindex+1))
            self.m.draw()
        
        
    def ClickNewCell(self):
        """ 
        this method is called when the button New Cell is clicked. If the button
        state corresponds to True (if is activated) then it connects the mouse 
        clicks on the pyqt window to the canvas (so to the matplolib figure).
        The connection has an "id" which is given by the integer self.id
        After the connections is made, it calls the Disable function with argument 0
        which turns off the other button(s).
        
        If the button is clicked but it is deactivated then it disconnects the 
        connection between the canvas and the window (the user can not interact
        with the plot anymore). 
        Storemouseclicks is a list corresponding to the coordinates of all mouse
        clicks between the activation and the deactivation of the button.
        So if it is empty, it does not draw anything because no clicks
        were registered. 
        But if it has some coordinates, it will draw a polygon where the vertices
        are the coordinates of all the mouseclicks.
        Once the figure has been updated with a new polygon, the other button(s)
        are again enabled. 
        
        """
        if self.button_newcell.isChecked():
            self.WriteStatusBar(('Draw a new cell. Use the left click to '
                                 'produce a polygon with a new cell value. '
                                 'Click the button again to confirm.'))
            self.m.tempmask = self.m.plotmask.copy()
            self.id = self.m.mpl_connect('button_press_event', self.m.MouseClick)
            self.Disable(self.button_newcell)
            
        else:
            self.m.mpl_disconnect(self.id)
            if  self.m.storemouseclicks and self.TestSelectedPoints():
                self.m.DrawRegion(True)
            else:
                self.m.updatedata()
            self.Enable(self.button_newcell)
            self.m.ShowCellNumbers()
            self.ClearStatusBar()
            
            
    def TestSelectedPoints(self):
        """This function is just used to catch an exception, when the new cell
        or the add region function is called. If all the dots drawn by the user
        are located on one line (horizontal or vertical) the DrawRegion 
        function calls a method to create a polygon and 
        it can not make a polygon out of straight line so it gives an error.
        In order to prevent this error, this function avoids to attempt to draw
        by returning False if the square are all on one line.
        """
                
        allx = list(np.array(self.m.storemouseclicks)[:,0])
        ally = list(np.array(self.m.storemouseclicks)[:,1])
        
        resultx = all(elem == allx[0] for elem in allx)
        resulty = all(elem == ally[0] for elem in ally)
        
        if resultx or resulty:
            return False
        else:
            return True
                    
        
    def clickmethod(self):
        """ 
        this method is called when the button Add region is clicked. If the button
        state corresponds to True (if is activated) then it connects the mouse 
        clicks on the pyqt window to the canvas (so to the matplolib figure).
        The connection has an "id" which is given by the integer self.id
        After the connections is made, it calls the Disable function with argument 1
        which turns off the other button(s).
        
        If the button is clicked and it is deactivated then it disconnects the 
        connection between the canvas and the window (the user can not interact
        with the plot anymore). 
        Storemouseclicks is a list corresponding to the coordinates of all mouse
        clicks between the activation and the deactivation of the button.
        So if it is empty, it does not draw anything because no clicks
        were registered. 
        But if it has some coordinates, it will draw a polygon where the vertices
        are the coordinates of all the mouseclicks.
        Once the figure has been updated with a new polygon, the other button(s)
        are again enabled. 
        
        """
        if self.button_add_region.isChecked():
            self.WriteStatusBar(('Select the cell with the first click, '
                                 'then draw polygon with subsequent '
                                 'clicks. Reclick the button to confirm.'))
            self.m.tempmask = self.m.plotmask.copy()
            self.id = self.m.mpl_connect('button_press_event', self.m.MouseClick)           
            self.Disable(self.button_add_region) 
            
        else:
            self.m.mpl_disconnect(self.id)
            
            # test if the list is not empty and if the dots are not all in the same line
            if self.m.storemouseclicks and self.TestSelectedPoints():
                self.m.DrawRegion(False)

            else:
                self.m.updatedata()
                
            self.Enable(self.button_add_region)
            self.m.ShowCellNumbers()
            self.ClearStatusBar()
            
            
# -----------------------------------------------------------------------------
# BUTTON ENABLE / DISABLE
    
    def Enable(self, button):
         """
         this functions turns on buttons all the buttons, depending on the time
         index. (next and previous buttons should not be turned on if t = 0 
         or t = lasttimeindex)
         """
         if self.button_segment.isChecked():
             self.EnableCorrectionsButtons()
             self.button_home.setEnabled(True)
             self.button_zoom.setEnabled(True)
             self.button_pan.setEnabled(True)
             self.button_back.setEnabled(True)
             self.button_forward.setEnabled(True)
         else:
             for k in range(0, len(self.buttonlist)):
                 if button != self.buttonlist[k]:
                     self.buttonlist[k].setEnabled(True)
                 
         if self.Tindex == 0:
             self.button_previousframe.setEnabled(False)
             
         if self.Tindex == self.reader.sizet-1:
             self.button_nextframe.setEnabled(False)
             
         self.EnableCNNButtons()
         
     
    def Disable(self, button):
         """
         this functions turns off all the buttons except the one given in 
         argument.
         """
         flag = False
         if (button == self.button_add_region 
             or button == self.button_newcell 
             or button == self.button_exval 
             or button == self.button_changecellvalue 
             or button == self.button_drawmouse 
             or button == self.button_eraser):
             if self.button_segment.isChecked():
                 flag = True

         for k in range(0,len(self.buttonlist)):
             if button != self.buttonlist[k]:
                 self.buttonlist[k].setEnabled(False)
         if flag:
             self.button_segment.setEnabled(True)
        
         if button == self.button_segment or button == self.button_threshold:
             self.button_home.setEnabled(True)
             self.button_zoom.setEnabled(True)
             self.button_pan.setEnabled(True)
             self.button_back.setEnabled(True)
             self.button_forward.setEnabled(True)


    def EnableCNNButtons(self):
        if self.reader.TestPredExisting(self.Tindex, self.FOVindex):
            self.button_threshold.setEnabled(True)
            self.button_segment.setEnabled(True)
            self.button_cellcorespondance.setEnabled(True)
            self.button_extractfluorescence.setEnabled(True)
        else:
            self.button_threshold.setEnabled(False)
            self.button_segment.setEnabled(False)
            self.button_cellcorespondance.setEnabled(False)
            self.button_extractfluorescence.setEnabled(False)
    
    
    def EnableCorrectionsButtons(self):
        self.button_newcell.setEnabled(True)
        self.button_add_region.setEnabled(True)
        self.button_drawmouse.setEnabled(True)
        self.button_eraser.setEnabled(True)
        self.button_exval.setEnabled(True)
        self.button_changecellvalue.setEnabled(True)
        self.button_showval.setEnabled(True)
        
        
    def DisableCorrectionsButtons(self):
        self.button_newcell.setEnabled(False)
        self.button_add_region.setEnabled(False)
        self.button_drawmouse.setEnabled(False)
        self.button_eraser.setEnabled(False)
        self.button_exval.setEnabled(False)
        self.button_changecellvalue.setEnabled(False)
        self.button_showval.setEnabled(False)
    
    
    def ButtonSaveMask(self):
        """
        When this function is called, it saves the current mask
        (self.m.plotmask)
        """
        self.reader.SaveMask(self.Tindex, self.FOVindex, self.m.plotmask)
        
        
    def WriteStatusBar(self, text):
        """Writes text to status bar"""
        self.statusBarText.setText(text)
        
        
    def ClearStatusBar(self):
        """Removes text from status bar"""
        self.statusBarText.setText('')


if __name__ == '__main__':
    app = QApplication(sys.argv)
    
    # If two arguments are given, make them nd2name and hdfname
    if len(sys.argv)==3:
        nd2name1 = sys.argv[1]
        hdfname1 = sys.argv[2]
        ex = App(nd2name1, hdfname1, '')
        sys.exit(app.exec_())
    
    # Launch file browser otherwise
    else:
        wind = dfb.FileBrowser()
        if wind.exec_():
            nd2name1 = wind.nd2name
            hdfname1 = wind.hdfname
            hdfnewname = wind.newhdfentry.text()
            ex = App(nd2name1, hdfname1, hdfnewname)
            sys.exit(app.exec_())
        else:
            app.exit()
        
